from argparse import ArgumentError
from logging.handlers import RotatingFileHandler
import logging as lg
import shutil, os, sys, time, openpyxl, zipfile
from openpyxl.styles import Border, Alignment, PatternFill, Font
from pathlib import Path
import pandas as pd


class Excel:

    changes_made = False
    backed_up = False
    ext_terminal = sys.stdout.isatty()

    def __init__(
        self,
        filename: str,
        use_logging: bool = True,
        log_file: str = "excel.log",
        log_level=lg.DEBUG,
    ):
        """
        Allows retreiving, adding, updating, deleting and
        formatting cells within Excel.

        `filename` is the path to the excel file.

        `use_logging` allows disabling all logs when running.

        `log_file` sets the path for logging.

        `log_level` Sets the logging level of this logger.
        level must be an int or a str.
        """
        # workbook setup
        self.file_path = Path(filename)
        try:
            self.wb = openpyxl.load_workbook(self.file_path)
        except zipfile.BadZipFile:
            print(f"Error with {self.file_path}.")
            response = input("Do you want to restore backup?")
            if response in ["yes", "yeah", "y"]:
                # renames current to .old
                os.rename(self.file_path, f"{self.file_path}.old")
                # renames backup to remove .bak
                os.rename(f"{self.file_path}.bak", self.file_path)
        # logger setup
        self.use_logging = use_logging
        datefmt = "%m-%d-%Y %I:%M:%S %p"
        log_formatter = lg.Formatter(
            "%(asctime)s %(levelname)s %(message)s", datefmt=datefmt
        )
        self.logger = lg.getLogger(__name__)
        self.logger.setLevel(log_level)  # Log Level
        max_gigs = 2
        my_handler = RotatingFileHandler(
            log_file,
            maxBytes=max_gigs * 1024 * 1024,
            backupCount=2,
        )
        my_handler.setFormatter(log_formatter)
        self.logger.addHandler(my_handler)

    def save(
        self,
        use_print: bool = True,
        force_save: bool = False,
        backup: bool = True,
    ):
        """
        `use_print` determines if info for the saving progress will be printed.

        `force_save` can be used to make sure a save occurs.

        Backs up the excel file before saving the changes if `backup` is True.

        It will keep trying to save until it completes in case of permission
        errors caused by the file being open.
        """
        # only saves if any changes were made
        if self.changes_made or force_save:
            try:
                # backups the file before saving.
                if backup:
                    if not self.backed_up:
                        backup_path = f"{self.file_path}.bak"
                        shutil.copy(self.file_path, backup_path)
                        self.backed_up = True
                        self.logger.info(f"Excel file backed up")
                # saves the file once it is closed
                if use_print:
                    print("\nSaving...")
                first_run = True
                while True:
                    try:
                        if self.file_path.exists:
                            self.wb.save(self.file_path)
                            if use_print:
                                print(f'Save Complete.{34*" "}')
                                self.changes_made = False
                            return True
                        else:
                            print("File no longer exists. Save Cancelled")
                        break
                    except PermissionError:
                        if first_run:
                            if use_print:
                                msg = "Make sure the excel sheet is closed."
                                print(msg, end="\r")
                            first_run = False
                        time.sleep(1)
            except KeyboardInterrupt:
                self.logger.warning("Save Cancelled")
                if use_print:
                    print("\nCancelling Save")
                exit()
        else:
            if use_print:
                msg = "Save Skipped due to no changes being made."
                self.logger.info(msg)
                print(msg)

    def open_excel(
        self,
        save: bool = True,
        exit_after: bool = True,
        test: bool = False,
    ):
        """
        Opens the current excel file if it still exists and then exits.

        Saves changes if `save` is True.
        """
        if save:
            self.save()
        if self.file_path.exists:
            if not test:
                os.startfile(self.file_path)
        else:
            # TODO raise Error
            print("File no longer exists.")
        if exit_after:
            exit()

    def open_file_input(self):
        """
        Opens the excel file if it exists after enter is
        pressed during the input.
        """
        if not self.ext_terminal:
            self.save()
            exit()
        try:
            input("\nPress Enter to open the excel sheet.\n")
        except KeyboardInterrupt:
            print("Closing...")
            self.save()
            exit()
        self.open_excel()


class Sheet:
    def __init__(
        self,
        excel_object: object,
        column_name: str,
        sheet_name: str = None,
        options: dict = None,
    ) -> None:
        """
        Allows interacting with any one sheet within the excel_object given.

        `excel_object` Excel object created using Excel class.

        `column_name` Name of the main column you intend to use for
        identifying rows.

        `sheet_name` Name of the sheet to use.

        `options` used to determine auto formatting.
        """
        self.wb = excel_object.wb
        self.excel = excel_object
        self.sheet_name = sheet_name
        self.column_name = column_name
        # defaults used sheet to first one if none is specified
        if sheet_name:
            self.cur_sheet = self.wb[sheet_name]
        else:
            if self.wb.sheetnames:
                self.cur_sheet = self.wb[self.wb.sheetnames[0]]
            else:
                raise "No sheets exist."
        self.column_name = column_name
        # column and row indexes
        self.col_idx = self.get_column_index()
        self.row_idx = self.get_row_index(self.column_name)
        # error checking
        self.missing_columns = []
        # formatting init
        # column format actions init
        self.column_formats = None
        # options
        self.options = options
        if not self.options:
            self.options = {
                "shrink_to_fit_cell": True,
                "fill": [],
                "percent": [
                    "%",
                    "Percent",
                ],
                "currency": ["Price", "MSRP", "Cost"],
                "integer": ["ID", "Number"],
                "count_days": ["Days Till", "Days Since"],
                "date": ["Last Updated", "Date"],
                "decimal": ["Hours"],
                "not_centered": ["Name"],
            }

    def create_dataframe(self, date_cols: list = None, na_vals: list = None):
        """
        Creates a panda dataframe using the current used sheet.

        `date_cols` sets the columns with dates.

        `na_vals` sets what should be considered N/A values that are ignored.
        """
        df = pd.read_excel(
            self.excel.file_path,
            engine="openpyxl",
            sheet_name=self.sheet_name,
            parse_dates=date_cols,
            na_values=na_vals,
        )
        return df

    @staticmethod
    def indirect_cell(left: int = 0, right: int = 0, manual_set: int = 0):
        """
        Returns a string for setting an indirect cell location to
        a number `left` or `right`.

        `manual_set` can be used to set the indirect cell offset manually.

        Only one direction can be greater then 0.
        """
        num = 0
        if left > 0 and right == 0:
            num -= left
        elif right > 0 and left == 0:
            num += right
        elif manual_set != 0:
            num = manual_set
        else:
            raise Exception("Left and Right args can't both be greater then 0.")
        return f'INDIRECT("RC[{num}]",0)'

    def easy_indirect_cell(self, cur_col: str, ref_col: str):
        """
        Allows setting up an indirect cell formula.

        Set `cur_col`to the column name of the column theformula is going
        into.

        Set `ref_col` to the column name of the column you are wanting
        to reference.
        """
        diff = self.col_idx[ref_col] - self.col_idx[cur_col]
        return self.indirect_cell(manual_set=diff)

    def get_column_index(self):
        """
        Creates the column index.
        """
        col_index = {}
        for i in range(1, len(self.cur_sheet["1"]) + 1):
            title = self.cur_sheet.cell(row=1, column=i).value
            if title is not None:
                col_index[title] = i
        return col_index

    def get_row_index(self, col_name: str):
        """
        Creates the row index based on `col_name`.
        """
        row_idx = {}
        total_rows = len(self.cur_sheet["A"])
        for row in range(1, total_rows):
            column = self.col_idx[col_name]
            title = self.cur_sheet.cell(row=row + 1, column=column).value
            if title is not None:
                row_idx[title] = row + 1
        return row_idx

    def list_in_string(self, list: list, string: str, lowercase: bool = True):
        """
        Returns True if any entry in the given `list` is in the given `string`.

        Setting `lowercase` to True allows you to make the check
        set all to lowercase.
        """
        if lowercase:
            return any(x.lower() in string.lower() for x in list)
        else:
            return any(x in string for x in list)

    def get_row_col_index(self, row_value: str or int, column_value: str or int):
        """
        Gets the row and column index for the given values if they exist.

        Will return the `row_value` and `column_value` if they are
        numbers already.
        """
        row_key, column_key = None, None
        # row key setup
        if type(row_value) == str and row_value in self.row_idx:
            row_key = self.row_idx[row_value]
        elif type(row_value) == int:
            row_key = row_value
        # column key setup
        if type(column_value) == str and column_value in self.col_idx:
            column_key = self.col_idx[column_value]
        elif type(column_value) == int:
            column_key = column_value
        return row_key, column_key

    def extract_hyperlink(self, cell_value):
        """
        Extracts the hyperlink target from a `cell_value` with the hyperlink
        formula.

        This is only needed if excel has not applied the hyperlink yet.
        This often happens when you click on the cell with the hyperlink
        formula.
        """
        if not cell_value:
            raise TypeError("Cell Value is None")
        if "=HYPERLINK(" in cell_value:
            split = cell_value.split('"')
            return split[1]
        else:
            raise ValueError("Cell Value is not an Excel hyperlink")

    def get_cell(self, row_value: str or int, column_value: str or int):
        """
        Gets the cell value based on the `row_value` and `column_value`.

        If the cell is a hyperlink that is currently clickable,
        the hyperlink target will be returned.
        """
        row_k, col_k = self.get_row_col_index(row_value, column_value)
        # gets the value
        if row_k is not None and col_k is not None:
            cell = self.cur_sheet.cell(row=row_k, column=col_k)
            if cell.hyperlink:
                return cell.hyperlink.target
            if type(cell.value) is str:
                # TODO add better regex test
                if "=HYPERLINK" in cell.value:
                    link = self.extract_hyperlink(cell.value)
                    if link:
                        return link
            return self.cur_sheet.cell(row=row_k, column=col_k).value
        else:
            return None

    def update_index(self, column_key: str):
        """
        Updates the current row with the `column_key` in the row_idx variable.
        """
        # TODO add test for this
        self.row_idx[column_key] = self.cur_sheet._current_row

    def update_cell(
        self,
        row_val: str,
        col_val: str,
        new_val: str or int,
        replace: bool = True,
        save: bool = False,
    ):
        """
        Updates the cell based on `row_val` and `col_val` to `new_val`.

        Returns True if cell was updated and False if it was not updated.

        `replace` allows you to determine if a cell will have its
        existing value changed if it is not None.

        Saves after change if `save` is True.
        """
        row_key, col_key = self.get_row_col_index(row_val, col_val)
        if row_key is not None and col_key is not None:
            cell = self.cur_sheet.cell(row=row_key, column=col_key)
            cur_val = cell.value
            # returns False if replace is False and the current value is not none
            if not replace and cur_val:
                return False
            # updates only if cell will actually be changed
            if new_val == "":
                new_val = None
            if cur_val != new_val:
                # FIXME datetime objects cause issues with this
                if cell.is_date:
                    pass
                self.cur_sheet.cell(row=row_key, column=col_key).value = new_val
                if save:
                    self.excel.save(use_print=False, backup=False)
                else:
                    self.excel.changes_made = True
                return True
        else:
            return False

    def add_new_line(
        self,
        cell_dict: dict,
        save: bool = False,
    ):
        """
        Adds cell_dict onto a new line within the excel sheet.
        The column_name must be given a value.

        If dictionary keys match existing columns within the set sheet,
        it will add the value to that column.

        Use `debug` to print info if a column in the `cell_dict` does not exist.

        Saves after change if `save` is True.
        """
        # missing column checker
        for col in cell_dict.keys():
            # TODO decide if this is needed
            if col not in self.col_idx and col not in self.missing_columns:
                self.missing_columns.append(col)
                msg = f"add_new_line: Missing {col} in {self.sheet_name} sheet"
                self.excel.log(msg, "warning")
        column_key = None
        append_list = []
        for col in self.col_idx:
            if col in cell_dict.keys():
                cell_value = cell_dict[col]
                # gets key for updating the index
                if self.column_name == col:
                    column_key = cell_value
            if col in cell_dict:
                append_list.append(cell_dict[col])
            else:
                append_list.append("")
        if not column_key:
            msg = "No Column given matches then sheets column key"
            raise ValueError(msg)
        self.cur_sheet.append(append_list)
        self.update_index(column_key)
        if save:
            self.excel.save(use_print=False, backup=False)
        else:
            self.excel.changes_made = True
        return True

    def delete_row(self, col_val: str, save: bool = False):
        """
        Deletes row by `column_value`.

        `save` allows you to force a save after deleting a row.
        """
        if col_val not in self.row_idx:
            return None
        row = self.row_idx[col_val]
        self.cur_sheet.delete_rows(row)
        self.row_idx.pop(col_val)  # removes index of row from row_idx
        if save:
            self.excel.save(use_print=False, backup=False)
        else:
            self.excel.changes_made = True
        return True

    def delete_column(self, column_name: str, save: bool = False):
        """
        Deletes column by `column_name`.
        """
        if column_name not in self.col_idx:
            return None
        column = self.col_idx[column_name]
        self.cur_sheet.delete_cols(column)
        if save:
            self.excel.save(use_print=False, backup=False)
        else:
            self.excel.changes_made = True
        return True

    # formatting

    def set_border(self, cell: object, style: str = "thin"):
        """
        Sets the given `cell` border to cover all sides with the given `style`.
        """
        cell.border = Border(
            left=openpyxl.styles.Side(style=style),
            right=openpyxl.styles.Side(style=style),
            top=openpyxl.styles.Side(style=style),
            bottom=openpyxl.styles.Side(style=style),
            outline=True,
        )

    def set_fill(
        self,
        cell: object,
        color: str = "000000",
        fill_type: str = "solid",
    ):
        """
        Sets the given `cell` to have fill with `color` and `fill_type`
        """
        cell.fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type=fill_type,
        )

    def set_style(self, cell: object, format: str = "general"):
        """
        Sets the given `cell` to the given `format` or general by default.
        """
        if format == "percent":
            cell.style = "Percent"
        elif format == "currency":
            cell.style = "Currency"
        else:
            cell.style = "General"

    def format_picker(self, column: str):
        """
        Determines what formatting to apply to a column.
        """
        option_keys = self.options.keys()
        actions = []
        # border
        actions.append("default_border")
        # alignment
        alignment = None
        if "default_align" in option_keys:
            alignment = self.options["default_align"]
        if "left_align" in option_keys:
            if column in self.options["left_align"]:
                alignment = "left_align"
            else:
                alignment = "center_align"
        if "right_align" in option_keys:
            if column in self.options["right_align"]:
                alignment = "right_align"
            else:
                alignment = "center_align"
        if alignment:
            actions.append(alignment)
        # fill
        if "black_fill" in option_keys:
            if self.list_in_string(self.options["black_fill"], column):
                actions.append("black_fill")
        elif "light_grey_fill" in option_keys:
            if self.list_in_string(self.options["light_grey_fill"], column):
                actions.append("light_grey_fill")
        # percent
        if "percent" in option_keys:
            if self.list_in_string(self.options["percent"], column):
                actions.append("percent")
                return actions
        # currency
        if "currency" in option_keys:
            if self.list_in_string(self.options["currency"], column):
                actions.append("currency")
                return actions
        if "integer" in option_keys:
            if column in self.options["integer"]:
                actions.append("integer")
                return actions
        # decimal
        # TODO allow variable decimal place
        if "decimal" in option_keys:
            if column in self.options["decimal"]:
                actions.append("decimal")
                return actions
        # countdown
        if "count_days" in option_keys:
            if column in self.options["count_days"]:
                actions.append("count_days")
                return actions
        # dates
        if "date" in option_keys:
            if self.list_in_string(self.options["date"], column):
                actions.append("date")
                return actions
        return actions

    def get_column_formats(self):
        """
        Gets the formats to use for each column.
        """
        format_actions = {}
        for column in self.col_idx.keys():
            actions = self.format_picker(column)
            if column not in format_actions.keys():
                format_actions[column] = actions
        return format_actions

    def format_header(self):
        """
        Formats the top header of the sheet.
        """
        header_options = self.options["header"]
        font_size = header_options["font_size"]
        bold_font = header_options["bold"]
        for column in self.col_idx.keys():
            col_i = self.col_idx[column]
            cell = self.cur_sheet.cell(row=1, column=col_i)
            cell.font = Font(
                name="Calibri",
                size=font_size,
                bold=bold_font,
                # color="FF000000",
            )

    def format_cell(self, column: str, row_i: int, col_i: int):
        """
        Formats a cell based on the `column` name using `row_i` and `col_i`.
        """
        # TODO add test for this
        cell = self.cur_sheet.cell(row=row_i, column=col_i)
        # gets format_actions if it has not be set yet
        if not self.column_formats:
            self.column_formats = self.get_column_formats()
        formatting = self.column_formats[column]
        # percent
        if "percent" in formatting:
            self.set_style(cell, format="percent")
        # currency
        elif "currency" in formatting:
            self.set_style(cell, format="currency")
        # integer
        elif "integer" in formatting:
            cell.number_format = "0"
        # decimal
        elif "decimal" in formatting:
            # TODO add decimal increase/decrease
            cell.number_format = "#,#0.0"
        # countdown
        elif "count_days" in formatting:
            cell.number_format = '# "Days"'
        # dates
        elif cell.is_date:
            cell.number_format = "mm-dd-yy"
        # border
        if "default_border" in formatting:
            self.set_border(cell)
        # alignment
        if "left_align" in formatting:
            cell.alignment = Alignment(horizontal="left")
        elif "center_align" in formatting:
            cell.alignment = Alignment(horizontal="center")
        elif "right_align" in formatting:
            cell.alignment = Alignment(horizontal="right")
        # fill
        if "black_fill" in formatting:
            self.set_fill(cell, color="fffff")
        elif "light_grey_fill" in formatting:
            self.set_fill(cell, color="F2F2F2")
        # makes sure changes will be saved next time the file is saved.
        self.changes_made = True

    def format_row(self, row_identifier: str):
        """
        Formats the entire row by `row_identifier`
        """
        # TODO add test for this
        for column in self.col_idx.keys():
            row_i = self.row_idx[row_identifier]
            col_i = self.col_idx[column]
            self.format_cell(column, row_i, col_i)

    def format_all_cells(self):
        """
        Auto formats all cells.
        TODO check for a way to make it use openpyxl more
        """
        # TODO add test for this
        # return early if options is not valid
        if not self.options:
            return False
        self.format_header()
        for column in self.col_idx.keys():
            # runs through all cells in a column and runs the actions
            col_i = self.col_idx[column]
            for row_i in self.row_idx.values():
                self.format_cell(column, row_i, col_i)
