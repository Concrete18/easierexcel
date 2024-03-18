import openpyxl
from openpyxl.styles import NamedStyle
import pandas as pd
from easierexcel import Excel

from dataclasses import dataclass, fields


@dataclass
class Sheet:
    """
    Allows interacting with any one sheet within the excel_object given.

    `excel_object` Excel object created using Excel class.

    `column_name` Name of the main column you intend to use for
    identifying rows.

    `sheet_name` Name of the sheet to use.

    `options` used to determine auto formatting.
    """

    excel_object: openpyxl.Workbook
    column_name: str
    sheet_name: str = None
    options: dict = None  # TODO Switch to using a new class

    def __post_init__(self):
        self.wb: openpyxl.Workbook = self.excel_object.wb
        self.excel: Excel = self.excel_object
        self.sheet_name = self.sheet_name
        self.column_name = self.column_name
        # defaults used sheet to first one if none is specified
        if self.sheet_name:
            if self.sheet_name in self.wb.sheetnames:
                self.cur_sheet = self.wb[self.sheet_name]
            else:
                raise Exception(f"{self.sheet_name} sheet does not exist.")
        else:
            self.sheet_name = self.wb.sheetnames[0]
            self.cur_sheet = self.wb[self.wb.sheetnames[0]]
        self.column_name = self.column_name
        # column and row indexes
        self.col_idx = self.get_column_index()
        self.row_idx = self.get_row_index(self.column_name)
        # error checking
        self.missing_columns = []
        # formatting init
        # column format actions init
        self.column_formats = None
        # custom formats

        # full date
        if not "full_date_format" in self.wb.named_styles:
            full_date_format_style = openpyxl.styles.NamedStyle(
                name="full_date_format", number_format="dddd, mmmm d, yyyy"
            )
            self.wb.add_named_style(full_date_format_style)

        # integer with commas
        if not "comma_format" in self.wb.named_styles:
            comma_format = NamedStyle(name="comma_format", number_format="#,##0")
            self.wb.add_named_style(comma_format)

        if not self.options:
            self.options = {
                "shrink_to_fit_cell": True,
                "fill": [],
                "percent": [
                    "%",
                    "Percent",
                ],
                "currency": ["Price", "MSRP", "Cost"],
                "integer": ["ID", "Number"],
                "count_days": ["Days Till", "Days Since"],
                "date": ["Last Updated", "Date"],
                "decimal": ["Hours"],
                "not_centered": ["Name"],
            }

    def __repr__(self):
        string = "Sheet("
        for field in fields(self):
            string += f"\n  {field.name}: {getattr(self, field.name)}"
        string += "\n)"
        return string

    def create_dataframe(self, date_cols: list = None, na_vals: list = None):
        """
        Creates a panda dataframe using the current used sheet.

        `date_cols` sets the columns with dates.

        `na_vals` sets what should be considered N/A values that are ignored.
        """
        df = pd.read_excel(
            self.excel.file_path,
            engine="openpyxl",
            sheet_name=self.sheet_name,
            parse_dates=date_cols,
            na_values=na_vals,
        )
        return df

    def indirect_cell(
        self,
        cur_col: str = None,
        ref_col: str = None,
        left: int = 0,
        right: int = 0,
        manual_set: int = 0,
    ):
        """
        Returns a string for setting an indirect cell location to a cell.

        If you want the cell to be relative to column names then set `cur_col`
        to the column name the formula is going into and `ref_col` to the
        column name you are wanting to reference.

        If you know it is simply references a cell that is 3 to the right or
        left then just give `left` or `right` that value. Only one direction
        can be greater than 0.

        You can also use `manual_set` to set the indirect cell offset manually
        using a positive or negative number.
        """
        num = 0
        if ref_col and cur_col:
            num = self.col_idx[ref_col] - self.col_idx[cur_col]
        elif left > 0 and right == 0:
            num -= left
        elif right > 0 and left == 0:
            num += right
        elif manual_set != 0:
            num = manual_set
        else:
            raise Exception("Left and Right args can't both be greater then 0.")
        return f'INDIRECT("RC[{num}]",0)'

    def get_column_index(self):
        """
        Creates the column index.
        """
        col_index = {}
        for i in range(1, len(self.cur_sheet["1"]) + 1):
            title = self.cur_sheet.cell(row=1, column=i).value
            if title is not None:
                col_index[title] = i
        return col_index

    def get_row_index(self, col_name: str):
        """
        Creates the row index based on `col_name`.
        """
        row_idx = {}
        total_rows = len(self.cur_sheet["A"])
        for row in range(1, total_rows):
            column = self.col_idx[col_name]
            cell_value = self.cur_sheet.cell(row=row + 1, column=column).value
            if cell_value is not None:
                row_idx[str(cell_value)] = row + 1
        return row_idx

    def list_in_string(self, list: list, string: str, lowercase: bool = True):
        """
        Returns True if any entry in the given `list` is in the given `string`.

        Setting `lowercase` to True allows you to make the check
        set all to lowercase.
        """
        if lowercase:
            return any(x.lower() in string.lower() for x in list)
        else:
            return any(x in string for x in list)

    def get_row_col_index(self, row_value: str | int, column_value: str | int):
        """
        Gets the row and column index for the given values if they exist.

        Will return the `row_value` and `column_value` if they are
        numbers already.
        """
        row_key, column_key = None, None
        row_value = str(row_value)
        column_value = str(column_value)
        # row key setup
        if row_value in self.row_idx:
            row_key = self.row_idx[row_value]
        # column key setup
        if column_value in self.col_idx:
            column_key = self.col_idx[column_value]
        return row_key, column_key

    def extract_hyperlink(self, cell_value):
        """
        Extracts the hyperlink target from a `cell_value` with the hyperlink
        formula.

        This is only needed if excel has not applied the hyperlink yet.
        This often happens when you click on the cell with the hyperlink
        formula.
        """
        if not cell_value:
            raise TypeError("Cell Value is None")
        if "=HYPERLINK(" in cell_value:
            split = cell_value.split('"')
            return split[1]
        else:
            raise ValueError("Cell Value is not an Excel hyperlink")

    def get_cell_by_key(self, row_key: int, column_key: int):
        """
        Gets the cell value based on the `row_key` and `column_key`.
        This is basically the index in excel for columns and rows.

        If the cell is a hyperlink that is currently clickable,
        the hyperlink target will be returned.
        """
        if row_key is not None and column_key is not None:
            cell = self.cur_sheet.cell(row=row_key, column=column_key)
            if cell.hyperlink:
                return cell.hyperlink.target
            if type(cell.value) is str:
                if cell.value.startswith("=HYPERLINK"):
                    link = self.extract_hyperlink(cell.value)
                    if link:
                        return link
            return self.cur_sheet.cell(row=row_key, column=column_key).value
        else:
            return None

    def get_cell(self, row_value: str | int, column_value: str | int):
        """
        Gets the cell value based on the `row_value` and `column_value`.

        If the cell is a hyperlink that is currently clickable,
        the hyperlink target will be returned.
        """
        # sets int to str
        if type(row_value) is int:
            row_value = str(row_value)
        if type(column_value) is int:
            column_value = str(column_value)
        # get row and column keys
        row_key, column_key = self.get_row_col_index(row_value, column_value)
        # returns the cell value
        return self.get_cell_by_key(row_key, column_key)

    def get_row(self, row_value: str | int):
        """
        Gets the row value based on the `row_value`.
        """
        # sets int to str
        row_value = str(row_value) if isinstance(row_value, int) else row_value
        # gets row dict
        row_dict = {}
        if row_value not in self.row_idx.keys():
            row_dict = {column: None for column in self.col_idx.keys()}
        else:
            row_data = self.cur_sheet[self.row_idx[row_value]]
            columns = list(self.col_idx.keys())
            row_dict = {col: entry.value for col, entry in zip(columns, row_data)}
        return row_dict

    def update_index(self, column_key: str):
        """
        Updates the current row with the `column_key` in the row_idx variable.
        """
        # TODO add test for this
        self.row_idx[str(column_key)] = self.cur_sheet._current_row

    def update_cell_by_key(
        self,
        row_key: int,
        col_key: int,
        new_val: str | int,
        replace: bool = True,
    ):
        """
        Updates the cell based on `row_key` and `col_key` to `new_val`.
        This is basically the index in excel for columns and rows.

        Returns True if cell was updated and False if it was not updated.

        `replace` allows you to determine if a cell will have its
        existing value changed if it is not None.
        """
        if not row_key and not col_key:
            return False
        cell = self.cur_sheet.cell(row=row_key, column=col_key)
        cur_val = cell.value
        # returns False if replace is False and the current value is not none
        if not replace and cur_val:
            return False
        # updates only if cell will actually be changed
        if new_val == "":
            new_val = None
        if cur_val != new_val:
            # FIXME datetime objects cause issues with this
            if cell.is_date:
                pass
            self.cur_sheet.cell(row=row_key, column=col_key).value = new_val
            self.excel.changes_made = True
            return True

    def update_cell(
        self,
        row_val: str,
        col_val: str,
        new_val: str | int,
        replace: bool = True,
    ):
        """
        Updates the cell based on `row_val` and `col_val` to `new_val`.

        Returns True if cell was updated and False if it was not updated.

        `replace` allows you to determine if a cell will have its
        existing value changed if it is not None.
        """
        row_key, col_key = self.get_row_col_index(row_val, col_val)
        return self.update_cell_by_key(row_key, col_key, new_val, replace)

    def clear_cell(
        self,
        row_val: str,
        col_val: str,
    ):
        """
        Clears the value of the cell based on `row_val` and `col_val`.
        """
        row_key, col_key = self.get_row_col_index(row_val, col_val)
        return self.update_cell_by_key(row_key, col_key, "", True)

    def add_new_line(self, cell_dict: dict):
        """
        Adds cell_dict onto a new line within the excel sheet.
        The column_name must be given a value.

        If dictionary keys match existing columns within the set sheet,
        it will add the value to that column.
        """
        column_key = None
        append_list = []
        for col in self.col_idx:
            if col in cell_dict.keys():
                cell_value = cell_dict[col]
                # gets key for updating the index
                if self.column_name == col:
                    column_key = cell_value
            if col in cell_dict:
                append_list.append(cell_dict[col])
            else:
                append_list.append("")
        if not column_key:
            msg = "No Column given matches the sheets column key"
            raise ValueError(msg)
        self.cur_sheet.append(append_list)
        self.update_index(column_key)
        self.excel.changes_made = True
        return True

    def delete_row(self, col_val: str):
        """
        Deletes row by `column_value`.
        """
        if col_val not in self.row_idx:
            return None
        row = self.row_idx[col_val]
        self.cur_sheet.delete_rows(row)
        self.row_idx.pop(col_val)  # removes index of row from row_idx
        self.row_idx = self.get_row_index(self.column_name)
        self.excel.changes_made = True
        return True

    def delete_column(self, column_name: str):
        """
        Deletes column by `column_name`.
        """
        if column_name not in self.col_idx:
            return None
        column = self.col_idx[column_name]
        self.cur_sheet.delete_cols(column)
        self.col_idx = self.get_column_index()
        self.excel.changes_made = True
        return True

    # formatting

    def set_border(self, cell, style: str = "thin"):
        """
        Sets the given `cell` border to cover all sides with the given `style`.
        """
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style=style),
            right=openpyxl.styles.Side(style=style),
            top=openpyxl.styles.Side(style=style),
            bottom=openpyxl.styles.Side(style=style),
            outline=True,
        )

    def set_fill(
        self,
        cell: object,
        color: str = "000000",
        fill_type: str = "solid",
    ):
        """
        Sets the given `cell` to have fill with `color` and `fill_type`
        """
        cell.fill = openpyxl.styles.PatternFill(
            start_color=color,
            end_color=color,
            fill_type=fill_type,
        )

    def set_style(self, cell: object, format: str = "general"):
        """
        Sets the given `cell` to the given `format` or general by default.
        """
        match format:
            case "percent":
                cell.style = "Percent"
            case "currency":
                cell.style = "Currency"
            case _:
                cell.style = "General"

    def get_style(self, row_value: str | int, column_value: str | int):
        """
        Gets the cell value based on the `row_value` and `column_value`.

        If the cell is a hyperlink that is currently clickable,
        the hyperlink target will be returned.
        """
        # sets int to str
        if type(row_value) is int:
            row_value = str(row_value)
        if type(column_value) is int:
            column_value = str(column_value)
        # get row and column keys
        row_key, column_key = self.get_row_col_index(row_value, column_value)
        # returns the cell style
        if row_key is not None and column_key is not None:
            cell = self.cur_sheet.cell(row=row_key, column=column_key)
            return cell.style

    def get_format(self, row_value: str | int, column_value: str | int):
        """
        Gets the cell value based on the `row_value` and `column_value`.

        If the cell is a hyperlink that is currently clickable,
        the hyperlink target will be returned.
        """
        # sets int to str
        if type(row_value) is int:
            row_value = str(row_value)
        if type(column_value) is int:
            column_value = str(column_value)
        # get row and column keys
        row_key, column_key = self.get_row_col_index(row_value, column_value)
        # returns the cell style
        if row_key is not None and column_key is not None:
            cell = self.cur_sheet.cell(row=row_key, column=column_key)
            return cell.style

    def format_picker(self, column: str):
        """
        Determines what formatting to apply to a column.
        """
        option_keys = self.options.keys()
        actions = []
        # border
        actions.append("default_border")
        # alignment
        alignment = None
        if "default_align" in option_keys:
            alignment = self.options["default_align"]
        if "left_align" in option_keys:
            if column in self.options["left_align"]:
                alignment = "left_align"
            else:
                alignment = "center_align"
        if "right_align" in option_keys:
            if column in self.options["right_align"]:
                alignment = "right_align"
            else:
                alignment = "center_align"
        if alignment:
            actions.append(alignment)
        # fill
        if "black_fill" in option_keys:
            if self.list_in_string(self.options["black_fill"], column):
                actions.append("black_fill")
        elif "light_grey_fill" in option_keys:
            if self.list_in_string(self.options["light_grey_fill"], column):
                actions.append("light_grey_fill")
        # percent
        if "percent" in option_keys:
            if self.list_in_string(self.options["percent"], column):
                actions.append("percent")
                return actions
        # currency
        if "currency" in option_keys:
            if self.list_in_string(self.options["currency"], column):
                actions.append("currency")
                return actions
        if "integer" in option_keys:
            if column in self.options["integer"]:
                actions.append("integer")
                return actions
        # decimal
        # TODO allow variable decimal place
        if "decimal" in option_keys:
            if column in self.options["decimal"]:
                actions.append("decimal")
                return actions
        # countdown
        if "count_days" in option_keys:
            if column in self.options["count_days"]:
                actions.append("count_days")
                return actions
        # dates
        if "full_date" in option_keys:
            if self.list_in_string(self.options["full_date"], column):
                actions.append("full_date")
                return actions
        elif "date" in option_keys:
            if self.list_in_string(self.options["date"], column):
                actions.append("date")
                return actions
        return actions

    def get_column_formats(self):
        """
        Gets the formats to use for each column.
        """
        format_actions = {}
        for column in self.col_idx.keys():
            actions = self.format_picker(column)
            if column not in format_actions.keys():
                format_actions[column] = actions
        return format_actions

    def format_header(self):
        """
        Formats the top header of the sheet.
        """
        header_options = self.options["header"]
        font_size = header_options["font_size"]
        bold_font = header_options["bold"]
        for column in self.col_idx.keys():
            col_i = self.col_idx[column]
            cell = self.cur_sheet.cell(row=1, column=col_i)
            cell.font = openpyxl.styles.Font(
                name="Calibri",
                size=font_size,
                bold=bold_font,
                # color="FF000000",
            )

    def auto_size_columns(self, width_multiplier=1.23, set_height=None):
        """
        ph
        """
        for col_cells in self.cur_sheet.columns:
            max_col_len = max(len(str(cell.value)) for cell in col_cells)
            new_col_lett = openpyxl.utils.get_column_letter(col_cells[0].column)
            if max_col_len > 0:
                col_width = max_col_len * width_multiplier
                self.cur_sheet.column_dimensions[new_col_lett].width = col_width

    def format_cell(self, column: str, row_i: int, col_i: int):
        """
        Formats a cell based on the `column` name using `row_i` and `col_i`.
        """
        # TODO add test for this
        cell = self.cur_sheet.cell(row=row_i, column=col_i)
        # gets format_actions if it has not be set yet
        if not self.column_formats:
            self.column_formats = self.get_column_formats()
        formatting = self.column_formats[column]
        # percent
        if "percent" in formatting:
            self.set_style(cell, format="percent")
        # currency
        elif "currency" in formatting:
            self.set_style(cell, format="currency")
        # integer
        elif "integer" in formatting:
            cell.number_format = "0"
        elif "comma_format" in formatting:
            cell.style = "comma_format"
        # decimal
        elif "decimal" in formatting:
            # TODO add decimal increase/decrease
            cell.number_format = "#,#0.0"
        # countdown
        elif "count_days" in formatting:
            cell.number_format = '# "Days"'
        # dates
        elif "full_date" in formatting:
            print("setting style")
            cell.style = "full_date_format"
        elif cell.is_date:
            cell.number_format = "mm-dd-yy"
        # border
        if "default_border" in formatting:
            self.set_border(cell)
        # alignment
        if "left_align" in formatting:
            cell.alignment = openpyxl.styles.Alignment(horizontal="left")
        elif "center_align" in formatting:
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        elif "right_align" in formatting:
            cell.alignment = openpyxl.styles.Alignment(horizontal="right")
        # fill
        if "black_fill" in formatting:
            self.set_fill(cell, color="fffff")
        elif "light_grey_fill" in formatting:
            self.set_fill(cell, color="F2F2F2")
        # makes sure changes will be saved next time the file is saved.
        self.changes_made = True

    def format_row(self, row_identifier: str):
        """
        Formats the entire row by `row_identifier`
        """
        # TODO add test for this
        row_i = self.row_idx[str(row_identifier)]
        for column in self.col_idx.keys():
            col_i = self.col_idx[column]
            self.format_cell(column, row_i, col_i)

    def format_all_cells(self):
        """
        Auto formats all cells.
        TODO check for a way to make it use openpyxl more
        """
        # TODO add test for this
        # return early if options is not valid
        if not self.options:
            return False
        self.format_header()
        for column in self.col_idx.keys():
            # runs through all cells in a column and runs the actions
            col_i = self.col_idx[column]
            for row_i in self.row_idx.values():
                self.format_cell(column, row_i, col_i)


if __name__ == "__main__":
    excel_obj = Excel(filename="test/excel_test.xlsx")
    sheet1 = Sheet(excel_obj, "Name")
    print(sheet1)
