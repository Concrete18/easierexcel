from logging.handlers import RotatingFileHandler
import logging as lg
import shutil, os, sys, time, openpyxl, zipfile
from openpyxl.styles import Border, Alignment, PatternFill, Font
from pathlib import Path
import pandas as pd


class Excel:

    changes_made = False
    backed_up = False
    ext_terminal = sys.stdout.isatty()

    def __init__(
        self,
        excel_filename,
        use_logging=True,
        log_file="excel.log",
        log_level=lg.DEBUG,
    ):
        """
        Allows retreiving, adding, updating, deleting and formatting cells within Excel.

        `excel_filename` is the path to the excel file.

        `log_file` sets the path for logging.

        `log_level` Sets the logging level of this logger. level must be an int or a str.
        """
        # workbook setup
        self.file_path = Path(excel_filename)
        try:
            self.wb = openpyxl.load_workbook(self.file_path)
        except zipfile.BadZipFile:
            response = input(
                f"Error with {self.file_path}.\nCheck backup to restore backup."
            )
            if response in ["yes", "yeah", "y"]:
                # renames current to .old
                os.rename(self.file_path, f"{self.file_path}.old")
                # renames backup to remove .bak
                os.rename(f"{self.file_path}.bak", self.file_path)
        # logger setup
        self.use_logging = use_logging
        datefmt = "%m-%d-%Y %I:%M:%S %p"
        log_formatter = lg.Formatter(
            "%(asctime)s %(levelname)s %(message)s", datefmt=datefmt
        )
        self.logger = lg.getLogger(__name__)
        self.logger.setLevel(log_level)  # Log Level
        max_gigs = 2
        my_handler = RotatingFileHandler(
            log_file,
            maxBytes=max_gigs * 1024 * 1024,
            backupCount=2,
        )
        my_handler.setFormatter(log_formatter)
        self.logger.addHandler(my_handler)

    def log(self, msg, type="info"):
        """
        Logs `msg` with set `type` if `use_logging` is True.
        """
        if self.use_logging:
            if type == "info":
                self.logger.info(msg)
            if type == "warning":
                self.logger.warning(msg)
            if type == "error":
                self.logger.error(msg)

    def create_dataframe(self, date_columns=None, na_values=None):
        """
        Creates a panda dataframe using the current used sheet.
        """
        file_loc = self.file_path
        df = pd.read_excel(
            file_loc,
            engine="openpyxl",
            sheet_name=None,
            parse_dates=date_columns,
            na_values=na_values,
        )
        return df

    def save_excel(self, use_print=True, force_save=False, backup=True):
        """
        Backs up the excel file before saving the changes if `backup` is True.

        It will keep trying to save until it completes in case of permission errors caused by the file being open.

        `use_print` determines if info for the saving progress will be printed.
        """
        # only saves if any changes were made
        if self.changes_made or force_save:
            try:
                # backups the file before saving.
                if backup:
                    if not self.backed_up:
                        shutil.copy(self.file_path, Path(self.file_path.name + ".bak"))
                        self.backed_up = True
                        self.log(f"Excel file backed up", "info")
                # saves the file once it is closed
                if use_print:
                    print("\nSaving...")
                first_run = True
                while True:
                    try:
                        if self.file_path.exists:
                            self.wb.save(self.file_path)
                            if use_print:
                                print(f'Save Complete.{34*" "}')
                                self.changes_made = False
                        else:
                            print("File no longer exists. Save Cancelled")
                        break
                    except PermissionError:
                        if first_run:
                            if use_print:
                                print("Make sure the excel sheet is closed.", end="\r")
                            first_run = False
                        time.sleep(1)
            except KeyboardInterrupt:
                self.log(f"Save Cancelled", "warning")
                if use_print:
                    print("\nCancelling Save")
                exit()
        else:
            if use_print:
                msg = "Save Skipped due to no changes being made."
                self.logger.info(msg)
                print(msg)

    def open_excel(self, save=True):
        """
        Opens the current excel file if it still exists and then exits.

        Saves changes if `save` is True.
        """
        if save:
            self.save_excel()
        if self.file_path.exists:
            os.startfile(self.file_path)
        else:
            print("File no longer exists.")
        exit()

    def open_file_input(self):
        """
        Opens the excel file if it exists after enter is pressed during the input.
        """
        if not self.ext_terminal:
            self.save_excel()
            exit()
        try:
            input("\nPress Enter to open the excel sheet.\n")
        except KeyboardInterrupt:
            print("Closing...")
            self.save_excel()
            exit()
        self.open_excel()


class Sheet:
    def __init__(
        self, excel_object, column_name, sheet_name=None, options=None
    ) -> None:
        """
        Allows interacting with any one sheet within the excel_object given.

        `excel_object` Excel object created using Excel class.

        `sheet_name` Name of the sheet to use.

        `column_name` Name of the main column you intend to use for identifying rows.
        """
        self.wb = excel_object.wb
        self.excel = excel_object
        self.column_name = column_name
        self.sheet_name = sheet_name
        # defaults used sheet to first one if none is specified
        if sheet_name:
            self.cur_sheet = self.wb[sheet_name]
        else:
            if self.wb.sheetnames:
                self.cur_sheet = self.wb[self.wb.sheetnames[0]]
            else:
                raise "No sheets exist."
        self.column_name = column_name
        # column and row indexes
        self.col_idx = self.get_column_index()
        self.row_idx = self.get_row_index(self.column_name)
        # error checking
        self.missing_columns = []
        # formatting init
        # column format actions init
        self.column_formats = None
        # options
        self.options = options
        if not self.options:
            self.options = {
                "shrink_to_fit_cell": True,
                "fill": [],
                "percent": [
                    "%",
                    "Percent",
                ],
                "currency": ["Price", "MSRP", "Cost"],
                "integer": ["ID", "Number"],
                "count_days": ["Days Till", "Days Since"],
                "date": ["Last Updated", "Date"],
                "decimal": ["Hours"],
                "not_centered": ["Name"],
            }

    @staticmethod
    def indirect_cell(left=0, right=0, manual_set=0):
        """
        Returns a string for setting an indirect cell location to a number `left` or `right`.

        Only one direction can be greater then 0.
        """
        num = 0
        if left > 0 and right == 0:
            num -= left
        elif right > 0 and left == 0:
            num += right
        elif manual_set != 0:
            num = manual_set
        else:
            raise "Left and Right can't both be greater then 0."
        return f'INDIRECT("RC[{num}]",0)'

    def easy_indirect_cell(self, cur_col, ref_col):
        """
        Allows setting up an indirect cell formula.

        Set `cur_col`to the column name of the column the formula is going into.

        Set `near_col` to the column name of the column you are wanting to reference.
        """
        diff = self.col_idx[ref_col] - self.col_idx[cur_col]
        return self.indirect_cell(manual_set=diff)

    def get_column_index(self):
        """
        Creates the column index.
        """
        col_index = {}
        for i in range(1, len(self.cur_sheet["1"]) + 1):
            title = self.cur_sheet.cell(row=1, column=i).value
            if title is not None:
                col_index[title] = i
        return col_index

    def get_row_index(self, col_name):
        """
        Creates the row index based on `column_name`.
        """
        row_idx = {}
        total_rows = len(self.cur_sheet["A"])
        for i in range(1, total_rows):
            title = self.cur_sheet.cell(row=i + 1, column=self.col_idx[col_name]).value
            if title is not None:
                row_idx[title] = i + 1
        return row_idx

    def list_in_string(self, list, string, lowercase=True):
        """
        Returns True if any entry in the given `list` is in the given `string`.

        Setting `lowercase` to True allows you to make the check set all to lowercase.
        """
        if lowercase:
            return any(x.lower() in string.lower() for x in list)
        else:
            return any(x in string for x in list)

    def get_row_col_index(self, row_value, column_value):
        """
        Gets the row and column index for the given values if they exist.

        Will return the `row_value` and `column_value` if they are numbers already.
        """
        row_key, column_key = None, None
        # row key setup
        if type(row_value) == str and row_value in self.row_idx:
            row_key = self.row_idx[row_value]
        elif type(row_value) == int:
            row_key = row_value
        # column key setup
        if type(column_value) == str and column_value in self.col_idx:
            column_key = self.col_idx[column_value]
        elif type(column_value) == int:
            column_key = column_value
        return row_key, column_key

    def get_cell(self, row_value, column_value):
        """
        Gets the cell value based on the `row_value` and `column_value`.
        """
        row_key, column_key = self.get_row_col_index(row_value, column_value)
        # gets the value
        if row_key is not None and column_key is not None:
            return self.cur_sheet.cell(row=row_key, column=column_key).value
        else:
            return None

    def update_index(self, col_key):
        """
        Updates the current row with the `column_key` in the `row_idx` variable.
        """
        self.row_idx[col_key] = self.cur_sheet._current_row

    def update_cell(self, row_val, col_val, new_val, replace=True, save=False):
        """
        Updates the cell based on `row_val` and `col_val` to `new_val`.

        Returns True if cell was updated and False if it was not updated.

        Saves after change if `save` is True.
        """
        row_key, col_key = self.get_row_col_index(row_val, col_val)
        if row_key is not None and col_key is not None:
            cur_val = self.cur_sheet.cell(row=row_key, column=col_key).value
            # returns False if replace is False and the current value is not none
            if not replace and cur_val is not None:
                return False
            # updates only if cell will actually be changed
            if new_val == "":
                new_val = None
            if cur_val != new_val:
                self.cur_sheet.cell(row=row_key, column=col_key).value = new_val
                if save:
                    self.excel.save_excel(use_print=False, backup=False)
                else:
                    self.excel.changes_made = True
                return True
        else:
            return False

    def add_new_line(self, cell_dict, column_key, save=False):
        """
        Adds the given dictionary, as `cell_dict`, onto a new line within the excel sheet.

        If dictionary keys match existing columns within the set sheet, it will add the value to that column.

        Use `debug` to print info if a column in the `cell_dict` does not exist.

        Saves after change if `save` is True.
        """
        # missing column checker
        for column in cell_dict.keys():
            if column not in self.col_idx and column not in self.missing_columns:
                self.missing_columns.append(column)
                msg = f"add_new_line: Missing {column} in {self.sheet_name} sheet"
                self.excel.log(msg, "warning")
        append_list = []
        for column in self.col_idx:
            if column in cell_dict:
                append_list.append(cell_dict[column])
            else:
                append_list.append("")
        self.cur_sheet.append(append_list)
        self.update_index(column_key)
        if save:
            self.excel.save_excel(use_print=False, backup=False)
        else:
            self.excel.changes_made = True
        return True

    def delete_by_row(self, col_val, save=False):
        """
        Deletes row by `column_value`.
        """
        if col_val not in self.row_idx:
            return None
        row = self.row_idx[col_val]
        self.cur_sheet.delete_rows(row)
        if save:
            self.excel.save_excel(use_print=False, backup=False)
        else:
            self.excel.changes_made = True
        return True

    def delete_by_column(self, column_name):
        """
        Deletes column by `column_name`.
        """
        if column_name not in self.col_idx:
            return None
        column = self.col_idx[column_name]
        self.cur_sheet.delete_column(column)
        self.excel.changes_made = True
        return True

    # formatting

    def set_border(self, cell, style="thin"):
        """
        ph
        """
        cell.border = Border(
            left=openpyxl.styles.Side(style=style),
            right=openpyxl.styles.Side(style=style),
            top=openpyxl.styles.Side(style=style),
            bottom=openpyxl.styles.Side(style=style),
            outline=True,
        )

    def set_fill(self, cell, color="000000", fill_type="solid"):
        """
        ph
        """
        cell.fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type=fill_type,
        )

    def set_style(self, cell, format="general"):
        """
        ph
        """
        if format == "percent":
            cell.style = "Percent"
        elif format == "currency":
            cell.style = "Currency"
        else:
            cell.style = "General"

    def set_date_format(self, cell, format="MM/DD/YYYY"):
        """
        Sets a cell to a date format.
        """
        if format == "MM/DD/YYYY":
            cell.number_format = "MM/DD/YYYY"

    def format_picker(self, column):
        """
        Determines what formatting to apply to a column.
        """
        actions = []
        # border
        actions.append("default_border")
        # alignment
        alignment = None
        if "left_align" in self.options.keys():
            if column in self.options["left_align"]:
                alignment = "left_align"
            else:
                alignment = "center_align"
        if "right_align" in self.options.keys():
            if column in self.options["right_align"]:
                actions.append("right_align")
            else:
                actions.append("center_align")
        if alignment:
            actions.append(alignment)
        # fill
        if "black_fill" in self.options.keys():
            if self.list_in_string(self.options["black_fill"], column):
                actions.append("black_fill")
        elif "light_grey_fill" in self.options.keys():
            if self.list_in_string(self.options["light_grey_fill"], column):
                actions.append("light_grey_fill")
        # percent
        if "percent" in self.options.keys():
            if self.list_in_string(self.options["percent"], column):
                actions.append("percent")
                return actions
        # currency
        if "currency" in self.options.keys():
            if self.list_in_string(self.options["currency"], column):
                actions.append("currency")
                return actions
        if "integer" in self.options.keys():
            if column in self.options["integer"]:
                actions.append("integer")
                return actions
        # decimal
        # TODO allow variable decimal place
        if "decimal" in self.options.keys():
            if column in self.options["decimal"]:
                actions.append("decimal")
                return actions
        # countdown
        if "count_days" in self.options.keys():
            if column in self.options["count_days"]:
                actions.append("count_days")
                return actions
        # dates
        if "date" in self.options.keys():
            if self.list_in_string(self.options["date"], column):
                actions.append("date")
                return actions
        return actions

    def get_column_formats(self):
        """
        Gets the formats to use for each column.
        """
        format_actions = {}
        for column in self.col_idx.keys():
            actions = self.format_picker(column)
            if column not in format_actions.keys():
                format_actions[column] = actions
        return format_actions

    def format_header(self):
        """
        ph
        """
        # TODO finish format_header function
        header_options = self.options["header"]
        font_size = header_options["font_size"]
        bold_font = header_options["bold"]
        for column in self.col_idx.keys():
            col_i = self.col_idx[column]
            cell = self.cur_sheet.cell(row=1, column=col_i)
            cell.font = Font(
                name="Calibri",
                size=font_size,
                bold=bold_font,
                # color="FF000000",
            )

    def format_cell(self, column, row_i, col_i):
        """
        ph
        """
        cell = self.cur_sheet.cell(row=row_i, column=col_i)
        # gets format_actions if it has not be set yet
        if not self.column_formats:
            self.column_formats = self.get_column_formats()
        formatting = self.column_formats[column]
        # percent
        if "percent" in formatting:
            self.set_style(cell, format="percent")
        # currency
        elif "currency" in formatting:
            self.set_style(cell, format="currency")
        # integer
        elif "integer" in formatting:
            cell.number_format = "0"
        # decimal
        elif "decimal" in formatting:
            # TODO add decimal increase/decrease
            cell.number_format = "#,#0.0"
        # countdown
        elif "count_days" in formatting:
            cell.number_format = '# "Days"'
        # dates
        elif "default_border" in formatting:
            self.set_date_format(cell, "")
        # border
        if "default_border" in formatting:
            self.set_border(cell)
        # alignment
        if "left_align" in formatting:
            cell.alignment = Alignment(horizontal="left")
        elif "center_align" in formatting:
            cell.alignment = Alignment(horizontal="center")
        elif "right_align" in formatting:
            cell.alignment = Alignment(horizontal="right")
        # fill
        if "black_fill" in formatting:
            self.set_fill(cell, color="fffff")
        elif "light_grey_fill" in formatting:
            self.set_fill(cell, color="F2F2F2")

    def format_row(self, row_identifier):
        """
        ph
        """
        for column in self.col_idx.keys():
            row_i = self.row_idx[row_identifier]
            col_i = self.col_idx[column]
            self.format_cell(column, row_i, col_i)

    def format_all_cells(self):
        """
        Auto formats all cells.
        """
        # return early if options is not valid
        if not self.options:
            return False
        self.format_header()
        for column in self.col_idx.keys():
            # runs through all cells in a column and runs the actions
            # TODO check for a way to make it use openpyxl more
            col_i = self.col_idx[column]
            for row_i in self.row_idx.values():
                cell = self.cur_sheet.cell(row=row_i, column=col_i)
                self.set_border(cell, "thick")
